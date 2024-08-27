
# Touchstone Feeder


import os
import json
from openpyxl import Workbook


# Get a list of all JSON files in the same folder as the script
dir_path = os.path.dirname(os.path.realpath(__file__))
json_files = [os.path.join(dir_path, f) for f in os.listdir(dir_path) if f.endswith('.json')]


# Create a new Excel workbook and sheet
wb = Workbook()
ws = wb.active


# Add column headers to the sheet
ws['A1'] = 'name'
ws['B1'] = 'description'
ws['C1'] = 'image'
ws['D1'] = 'supply'
ws['E1'] = 'rarity'
ws['F1'] = 'price'


# User inputs their legendary, rare, uncommon and common price, description and supply
legendary_price = int(input("Legendary price: "))
rare_price = int(input("Rare price: "))
uncommon_price = int(input("Uncommon price: "))
common_price = int(input("Common price: "))
mydescription = (input("Description: "))
mysupply = int(input("Supply: "))


# Function to get rarity and price attribute
def getrarity(data):


    # Default values for rarity and price - if this shows up on your excel sheet, something is wrong
    myrarity = "unknown"
    myprice = "unknown"


    with open(json_file, 'r') as f:
        metadata = json.load(f)


    # !IMPORTANT! Customize these traits to your NFTs. Example: "Layer": ["Trait", "Trait"]
    legendary_traits = {
        "Background": ["Day", "Night"],
        "Body": ["Rainbowbody"],
        "Tea": ["Rainbowtea"],
        "Bubble": ["Glowing"],
        "Face": ["Blueface", "Purpleface"],
        "Expression": ["Evil", "Hearteyes"],
        "Accessories": ["Cow", "Ufo"],
        "Straw": ["Shinystraw"]
    }


    rare_traits = {
        "Background": ["Blorangedots", "Preendots", "Yelangedots"],
        "Body": ["Blackgradient", "Orangegradient", "Turquoisegradient", "Yellowgradient"],
        "Tea": ["Bluegradienttea", "Browngradienttea", "Greengradienttea", "Pinkgradienttea", "Yellowgradienttea"],
        "Bubble": ["Rainbowbubble", "Heartblue"],
        "Face": ["Greenface", "Yellowface"],
        "Expression": ["Cat", "Dog", "Ghost"],
        "Accessories": ["Crown", "Headband", "Headphones"],
        "Straw": ["Neon", "Silverstraw"]
    }


    uncommon_traits = {
        "Background": ["Blueconfetti", "Greenconfetti", "Pinkconfetti", "Yellowconfetti"],
        "Body": ["Bellowbody", "Bledbody", "Gurplebody", "Orinkbody"],
        "Tea": ["Bluetea", "Browntea", "Greentea", "Pinktea", "Purpletea", "Redtea", "Yellowtea"],
        "Bubble": ["Bluebubble", "Goldbubble", "Greenbubble", "Lightbluebubble", "Orangebubble", "Pinkbubble", "Purpbubble", "Whitebubble", "Yellowbubble"],
        "Face": ["Orangeface", "Redface"],
        "Expression": ["Crying", "Sideways", "Silly", "Smirk"],
        "Accessories": ["Chain", "Glasses", "Sunglasses"],
        "Straw": ["Bluestraw", "Brownstraw", "Orangestraw", "Pinkstraw", "Purplestraw", "Redstraw", "Rosestraw", "Turquoisestraw", "Yelangestraw"]
    }


    common_traits = {
        "Background": ["Solidblue", "Solidgreen", "Solidpink", "Solidpurple", "Solidyellow"],
        "Body": ["Blackbody", "Greenbody", "Orangebody", "Turquoisebody", "Yellowbody"],
        "Tea": ["Beige"],
        "Bubble": ["Blackbody"],
        "Face": ["Skin", "Skin2", "Skin3", "Skin4", "Skin5", "Skin6", "Skin7", "Skin8"],
        "Expression": ["Annoyed", "Bruh", "Eyelash", "Lazy", "Smile"],
        "Accessories": ["Bow", "Bowtie", "Monocle", "Moustache", "Party_hat"],
        "Straw": ["Darkplain", "Plain"]
    }


    # Checks if legendary, rare, uncommon or common trait is inside metadata


    legendary = any(
        trait["value"] in legendary_traits.get(trait["trait_type"], [])
        for trait in metadata.get("attributes", [])
    )


    rare = any(
        trait["value"] in rare_traits.get(trait["trait_type"], [])
        for trait in metadata.get("attributes", [])
    )


    uncommon = any(
        trait["value"] in uncommon_traits.get(trait["trait_type"], [])
        for trait in metadata.get("attributes", [])
    )


    common = any(
        trait["value"] in rare_traits.get(trait["trait_type"], [])
        for trait in metadata.get("attributes", [])
    )
   
    # Sets rarity and price attribute to corresponding category
    if legendary is True:
        myrarity = "legendary"
        myprice = legendary_price
    elif rare is True:
        myrarity = "rare"
        myprice = rare_price
    elif uncommon is True:
        myrarity = "uncommon"
        myprice = uncommon_price
    else:
        myrarity = "common"
        myprice = common_price


    return myrarity, myprice


# Iterate over each JSON file and extract attributes.
for i, json_file in enumerate(json_files):
    with open(json_file, 'r') as f:
        data = json.load(f)
        name = data.get('name', '')
        description = mydescription
        image = name + ".png"
        supply = mysupply
        myrarity, myprice = getrarity(data)  # <-- call getrarity and pass in data as an argument
        rarity = myrarity
        price = myprice


        # Write attributes to the sheet
        ws.cell(row=i+2, column=1, value=name)
        ws.cell(row=i+2, column=2, value=description)
        ws.cell(row=i+2, column=3, value=image)
        ws.cell(row=i+2, column=4, value=supply)
        ws.cell(row=i+2, column=5, value=rarity)
        ws.cell(row=i+2, column=6, value=price)


# Save the workbook as an Excel file
wb.save('descriptions.xlsx')


